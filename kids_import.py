import openpyxl
import mysql.connector
import re

# === CONFIGURATION ===
EXCEL_FILE = "kids.xlsx"
TABELLENBLATT = "dbo_Tourkinder"
DB = dict(
    host="localhost",
    user="usr_tk_ctrl8",
    password="SafeTaxi2025_",
    database="tekath_control",
    charset="utf8mb4"
)

MAPPING = {
    "TKNachname": "last_name",
    "TKVorname": "first_name",
    "TKStrasse": "street",
    "TKHeimPLZ": "zip",
    "TKHeimOrt": "city",
    "TKTelefon": "kid_phone",
    "TKFahrziel": "school_shorttag",
}

def split_street(street):
    # Trennt Straße und Hausnummer ("Musterstraße 123a" → "Musterstraße", "123a")
    if not street:
        return "", ""
    match = re.match(r"^(.+?)\s+(\d+[a-zA-Z\-]*)\s*$", str(street).strip())
    if match:
        return match.group(1).strip(), match.group(2).strip()
    else:
        # Wenn keine Hausnummer gefunden wird, alles als Straße, Hausnummer leer
        return str(street).strip(), ""

def get_column_indices(header, mapping):
    idx = {}
    for col, dest in mapping.items():
        try:
            idx[dest] = header.index(col)
        except ValueError:
            print(f"⚠️ Excel-Spalte '{col}' nicht gefunden!")
    return idx

def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    if TABELLENBLATT not in wb.sheetnames:
        print(f"❌ Blatt '{TABELLENBLATT}' nicht gefunden! Vorhandene Blätter: {wb.sheetnames}")
        return
    sheet = wb[TABELLENBLATT]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    idx = get_column_indices(header, MAPPING)
    
    conn = mysql.connector.connect(**DB)
    cursor = conn.cursor(dictionary=True)

    count_inserted = 0
    count_skipped = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            last_name   = row[idx["last_name"]]
            first_name  = row[idx["first_name"]]
            street_raw  = row[idx["street"]]
            zip_code    = str(row[idx["zip"]]) if row[idx["zip"]] else None
            city        = row[idx["city"]]
            kid_phone   = row[idx["kid_phone"]]
            shorttag    = row[idx["school_shorttag"]]
        except Exception as e:
            print(f"⚠️ Fehler beim Lesen der Zeile: {e}")
            count_skipped += 1
            continue

        # Straße und Hausnummer extrahieren
        street, street_number = split_street(street_raw)

        if not shorttag:
            print(f"⚠️ Keine Schule angegeben für {first_name} {last_name}. Überspringe…")
            count_skipped += 1
            continue

        try:
            cursor.execute("SELECT id FROM school WHERE shorttag = %s", (shorttag,))
            school = cursor.fetchone()
            while cursor.nextset():
                pass
        except Exception as ex:
            print(f"❌ Fehler beim Suchen der Schule für {first_name} {last_name}: {ex}")
            count_skipped += 1
            continue

        if not school:
            print(f"❌ Keine Schule gefunden für TKFahrziel '{shorttag}' (Schüler: {first_name} {last_name})")
            count_skipped += 1
            continue
        school_id = school['id']

        # Insert Schoolkid (JETZT MIT street_number!)
        sql = """
        INSERT INTO schoolkids
        (last_name, first_name, street, street_number, zip, city, kid_phone, school_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        values = (last_name, first_name, street, street_number, zip_code, city, kid_phone, school_id)
        try:
            cursor.execute(sql, values)
            print(f"✅ Importiert: {first_name} {last_name} ({shorttag} → School ID {school_id}, {street} {street_number})")
            count_inserted += 1
        except Exception as ex:
            print(f"❌ Fehler beim Import von {first_name} {last_name}: {ex}")
            count_skipped += 1

    conn.commit()
    cursor.close()
    conn.close()
    print(f"\nFERTIG! {count_inserted} Einträge importiert, {count_skipped} übersprungen.")

if __name__ == "__main__":
    main()
